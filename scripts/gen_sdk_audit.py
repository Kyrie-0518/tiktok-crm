from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
wb.remove(wb.active)

header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2563EB')
center_font = Font(name='Arial', size=10)
yes_fill = PatternFill('solid', fgColor='DCFCE7')
yes_font = Font(name='Arial', size=10, color='166534')
no_fill = PatternFill('solid', fgColor='FEE2E2')
no_font = Font(name='Arial', size=10, color='991B1B')
warn_fill = PatternFill('solid', fgColor='FEF3C7')
warn_font = Font(name='Arial', size=10, color='92400E')
thin_border = Border(
    left=Side(style='thin', color='E5E7EB'),
    right=Side(style='thin', color='E5E7EB'),
    top=Side(style='thin', color='E5E7EB'),
    bottom=Side(style='thin', color='E5E7EB'),
)
center_align = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

data = {
    '店铺中心': [
        ('店铺ID', '✅', 'ShopsGet / ShopsGetShops'),
        ('店铺名称', '✅', 'ShopsGet'),
        ('店铺类型（跨境/本土）', '✅', 'ShopsGet'),
        ('站点（US/UK等）', '✅', 'ShopsGet'),
        ('主营类目', '❌', '需产品列表反推'),
        ('开店时间', '✅', 'ShopsGet'),
        ('店铺评分', '❌', '未开放'),
        ('Seller Score', '❌', '未开放'),
        ('Shop Performance Score', '⚠️', 'ShopPerformanceGet（有指标但非分数值）'),
        ('店铺状态', '✅', 'ShopsGet'),
        ('违规记录', '❌', '未开放'),
        ('处罚记录', '❌', '未开放'),
        ('默认仓库', '⚠️', 'SupplierShopsGet / 单独接口'),
        ('物流模板', '⚠️', 'LogisticsV202309 单独接口'),
    ],
    '商品中心': [
        ('Product ID', '✅', 'SearchProducts / GetProduct'),
        ('SPU', '✅', 'global_product_id'),
        ('SKU', '✅', 'skus[].id / skus[].seller_sku'),
        ('商品标题', '✅', 'title'),
        ('商品副标题', '❌', '无独立字段'),
        ('品牌', '✅', 'brand_id + BrandsGet'),
        ('类目', '✅', 'category_id + GetCategories'),
        ('商品属性', '✅', 'product_attributes / GetAttributes'),
        ('商品标签', '⚠️', '平台标签在listing_check中'),
        ('主图', '✅', 'main_image / ImagesUploadPost'),
        ('图片集', '✅', 'images[]'),
        ('商品视频', '✅', 'video字段'),
        ('详情页', '✅', 'description'),
        ('卖点描述', '❌', '无独立字段'),
        ('规格参数', '✅', 'skus[].sales_attributes'),
        ('商品状态', '✅', 'status（ACTIVATE/DEACTIVATE）'),
        ('创建时间', '✅', 'create_time'),
    ],
    '价格中心': [
        ('原价', '✅', 'original_price'),
        ('售价', '✅', 'current_price'),
        ('SKU价格', '✅', 'skus[].price'),
        ('活动价', '⚠️', 'promotion接口获取'),
        ('达人价', '⚠️', 'affiliate接口'),
        ('会员价', '❌', '未开放'),
        ('组合(Bundle)价格', '❌', '未开放'),
        ('平台补贴', '⚠️', 'promotion间接获取'),
        ('Seller Coupon', '✅', 'CreateCoupon / GetCoupon'),
        ('Platform Coupon', '⚠️', '平台券部分曝光'),
        ('优惠券', '✅', 'CouponsGet / CouponsPost'),
        ('满减', '✅', 'coupon规则配置'),
        ('秒杀（Flash Sale）', '✅', 'FlashSalesGet / FlashSalesProducts'),
        ('包邮活动', '✅', 'FreeShippingGet / FreeShippingPost'),
        ('折扣比例', '⚠️', '需计算：原价-售价/原价'),
    ],
    '库存中心': [
        ('总库存', '✅', 'InventorySearchPost / UpdateGlobalInventory'),
        ('可售库存', '✅', 'available_quantity'),
        ('锁定库存', '✅', 'reserved_quantity'),
        ('安全库存', '❌', '未开放'),
        ('海外仓库存', '✅', 'warehouse_inventory分仓库返回'),
        ('跨境仓库存', '✅', '同上'),
        ('仓库名称', '✅', 'warehouse_id + 仓库接口查询'),
        ('库存更新时间', '✅', '响应时间戳'),
    ],
    '内容中心': [
        ('Video ID', '✅', '平台视频ID（商品关联视频）'),
        ('视频名称', '❌', '未开放'),
        ('视频封面', '⚠️', '商品视频有URL'),
        ('视频时长', '❌', '未开放'),
        ('比例', '❌', '未开放'),
        ('分辨率', '❌', '未开放'),
        ('发布时间', '❌', '未开放'),
        ('关联商品', '✅', '商品视频关联'),
        ('关联SKU', '❌', '未开放'),
        ('AI模型（Seedance等）', '❌', '虾掌柜自有'),
        ('Prompt', '❌', '虾掌柜自有'),
        ('优化Prompt', '❌', '虾掌柜自有'),
        ('Prompt版本', '❌', '虾掌柜自有'),
        ('导演方案', '❌', '虾掌柜自有'),
        ('镜头脚本', '❌', '虾掌柜自有'),
        ('视频风格', '❌', '虾掌柜自有'),
        ('运镜', '❌', '虾掌柜自有'),
        ('灯光', '❌', '虾掌柜自有'),
        ('色调', '❌', '虾掌柜自有'),
        ('Negative Prompt', '❌', '虾掌柜自有'),
        ('生成参数', '❌', '虾掌柜自有'),
        ('生成次数', '❌', '虾掌柜自有'),
        ('生成耗时', '❌', '虾掌柜自有'),
    ],
    '视频表现': [
        ('曝光（Impression）', '✅', 'video_impressions'),
        ('播放量（Views）', '✅', 'video_views'),
        ('3秒播放率', '⚠️', '部分版本video_3sec_views'),
        ('5秒播放率', '❌', '未直接暴露'),
        ('完播率', '⚠️', 'video_completion_rate'),
        ('平均观看时长', '✅', 'average_watch_time'),
        ('点赞', '✅', 'video_likes'),
        ('评论', '✅', 'video_comments'),
        ('收藏', '✅', 'video_favorites'),
        ('分享', '✅', 'video_shares'),
        ('商品点击', '✅', 'product_clicks'),
        ('CTR', '⚠️', 'product_clicks/impressions'),
        ('加购', '✅', 'add_to_cart'),
        ('CVR', '⚠️', 'orders/product_clicks'),
        ('订单数', '✅', 'orders'),
        ('GMV', '✅', 'gmv'),
        ('ROI', '❌', '需计算'),
        ('ROAS', '❌', '需计算'),
    ],
    '流量中心': [
        ('总曝光', '✅', 'impression（Shop/Product Performance）'),
        ('总点击', '✅', 'click'),
        ('CTR', '⚠️', '可计算'),
        ('PV', '✅', 'page_view'),
        ('UV', '✅', 'unique_visitor'),
        ('搜索流量', '⚠️', '部分版本search_traffic'),
        ('推荐流量', '⚠️', '归因类型分类'),
        ('商城流量', '⚠️', '同上'),
        ('达人流量', '✅', 'affiliate_traffic'),
        ('直播流量', '✅', 'live_traffic'),
        ('广告流量', '✅', 'ads_traffic'),
        ('外部流量', '⚠️', 'external_traffic'),
        ('自然流量', '⚠️', '总流量-广告-达人'),
        ('流量趋势（今日/7天/30天）', '✅', 'granularity（HOUR/DAY）聚合'),
    ],
    '广告中心': [
        ('Campaign ID', '✅', 'campaign_id'),
        ('Campaign名称', '✅', 'campaign_name'),
        ('Ad Group', '✅', 'adgroup_id / adgroup_name'),
        ('Creative', '✅', 'creative_id'),
        ('广告状态', '✅', 'status'),
        ('预算', '✅', 'budget'),
        ('花费', '✅', 'spend'),
        ('曝光', '✅', 'impressions'),
        ('点击', '✅', 'clicks'),
        ('CTR', '⚠️', '可计算'),
        ('CPM', '✅', 'cpm'),
        ('CPC', '✅', 'cpc'),
        ('CPA', '⚠️', '可计算'),
        ('CVR', '⚠️', '可计算'),
        ('订单数', '✅', 'orders'),
        ('GMV', '✅', 'gmv'),
        ('ROAS', '✅', 'roas'),
        ('ROI', '❌', '需计算'),
        ('Learning状态', '✅', 'learning_phase'),
        ('审核状态', '✅', 'review_status'),
        ('开始时间', '✅', 'start_time'),
        ('结束时间', '✅', 'end_time'),
    ],
    '订单中心': [
        ('订单ID', '✅', 'order_id'),
        ('订单状态', '✅', 'order_status'),
        ('Product ID', '✅', 'line_items[].product_id'),
        ('SKU', '✅', 'line_items[].sku_id / seller_sku'),
        ('购买数量', '✅', 'line_items[].quantity'),
        ('支付金额', '✅', 'payment_amount'),
        ('优惠金额', '✅', 'discount_amount'),
        ('运费', '✅', 'shipping_fee'),
        ('买家地区', '✅', 'buyer_region / 收货地址'),
        ('下单时间', '✅', 'create_time'),
        ('支付时间', '✅', 'payment_time'),
        ('发货时间', '✅', 'ship_time'),
        ('签收时间', '✅', 'delivery_time'),
        ('完成时间', '✅', 'complete_time'),
        ('退款状态', '✅', 'return_refund_status'),
        ('退款原因', '✅', 'return_reason'),
        ('取消原因', '✅', 'cancel_reason'),
    ],
    '物流中心': [
        ('仓库类型（海外仓/跨境仓）', '✅', 'warehouse_type'),
        ('物流方式', '✅', 'shipping_provider'),
        ('物流商', '✅', 'shipping_provider_id'),
        ('预计时效', '✅', 'estimated_delivery_time'),
        ('实际时效', '✅', 'actual_delivery_time'),
        ('平均发货时间', '✅', 'Analytics或自行聚合'),
        ('平均签收时间', '✅', '同上'),
        ('物流异常率', '⚠️', '需自行计算'),
        ('拒收率', '❌', '未直接暴露'),
        ('取消率', '⚠️', '从订单数据计算'),
        ('超时率', '❌', '未直接暴露'),
    ],
    '评价中心': [
        ('Review ID', '✅', 'review_id'),
        ('评分', '✅', 'rating（1-5）'),
        ('评论内容', '✅', 'content'),
        ('评论图片', '✅', 'images[]'),
        ('评论视频', '✅', 'videos[]'),
        ('评论时间', '✅', 'create_time'),
        ('高频关键词', '❌', '需自行分析'),
        ('情绪分析', '❌', '需自行分析'),
        ('好评率', '⚠️', '4-5星占比'),
        ('差评率', '⚠️', '1-2星占比'),
        ('用户反馈分类', '❌', '未开放'),
    ],
    '财务中心': [
        ('GMV', '✅', 'sales_gmv（Analytics）'),
        ('净销售额', '✅', 'net_sales'),
        ('广告花费', '✅', 'Ads API spend'),
        ('物流成本', '⚠️', 'Settlement接口间接获取'),
        ('商品成本', '❌', '虾掌柜自有'),
        ('平台佣金', '✅', 'platform_commission'),
        ('退款金额', '✅', 'refund_amount'),
        ('利润', '❌', '需计算'),
        ('利润率', '❌', '需计算'),
        ('ROI', '❌', '需计算'),
        ('ROAS', '✅', 'Ads API'),
        ('CAC（获客成本）', '❌', '需计算'),
        ('客单价', '✅', 'aov（Analytics）'),
    ],
    '事件中心': [
        ('修改标题', '❌', '未开放'),
        ('修改主图', '❌', '未开放'),
        ('修改详情页', '❌', '未开放'),
        ('修改价格', '❌', '未开放'),
        ('更换仓库', '❌', '未开放'),
        ('库存售罄', '⚠️', '可轮询检测'),
        ('恢复库存', '❌', '未开放'),
        ('开启广告', '⚠️', '可轮询检测'),
        ('关闭广告', '⚠️', '可轮询检测'),
        ('新增达人合作', '⚠️', 'Affiliate API变更可检测'),
        ('开启秒杀', '⚠️', 'Promotion API变更可检测'),
        ('开启优惠券', '⚠️', '同上'),
        ('修改Prompt', '❌', '虾掌柜自有'),
        ('重新生成视频', '❌', '虾掌柜自有'),
        ('修改视频', '❌', '虾掌柜自有'),
        ('API异常', '❌', '虾掌柜自有'),
        ('店铺授权失效', '❌', '虾掌柜自有'),
        ('库存同步失败', '❌', '虾掌柜自有'),
    ],
}

def style_status(cell, status):
    cell.font = Font(name='Arial', size=10)
    cell.border = thin_border
    cell.alignment = center_align
    if status == '✅':
        cell.fill = yes_fill
        cell.font = yes_font
    elif status == '❌':
        cell.fill = no_fill
        cell.font = no_font
    elif status == '⚠️':
        cell.fill = warn_fill
        cell.font = warn_font

for sheet_name, rows in data.items():
    ws = wb.create_sheet(title=sheet_name)
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 42

    ws.merge_cells('A1:D1')
    ws['A1'] = f'{sheet_name} — SDK字段审计'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='1E293B')
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 32

    headers = ['字段', '要/不要', 'SDK支持', '来源/接口']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[2].height = 28

    for row_idx, (field, status, source) in enumerate(rows, 3):
        ws.cell(row=row_idx, column=1, value=field).font = center_font
        ws.cell(row=row_idx, column=1).border = thin_border
        ws.cell(row=row_idx, column=1).alignment = left_align

        ws.cell(row=row_idx, column=2, value='').border = thin_border
        ws.cell(row=row_idx, column=2).alignment = center_align

        style_status(ws.cell(row=row_idx, column=3), status)
        ws.cell(row=row_idx, column=3).value = status

        ws.cell(row=row_idx, column=4, value=source).font = Font(name='Arial', size=9, color='6B7280')
        ws.cell(row=row_idx, column=4).border = thin_border
        ws.cell(row=row_idx, column=4).alignment = left_align

        ws.row_dimensions[row_idx].height = 22

ws_summary = wb.create_sheet(title='汇总', index=0)
ws_summary.column_dimensions['A'].width = 18
ws_summary.column_dimensions['B'].width = 11
ws_summary.column_dimensions['C'].width = 14
ws_summary.column_dimensions['D'].width = 14
ws_summary.column_dimensions['E'].width = 14
ws_summary.column_dimensions['F'].width = 14

ws_summary.merge_cells('A1:F1')
ws_summary['A1'] = 'TikTok Shop SDK 全量字段审计 — 虾掌柜ERP'
ws_summary['A1'].font = Font(name='Arial', bold=True, size=14, color='1E293B')
ws_summary.row_dimensions[1].height = 36

summary_headers = ['数据中心', '总数', 'SDK支持✅', '可计算⚠️', '不支持❌', '覆盖率']
for col_idx, h in enumerate(summary_headers, 1):
    cell = ws_summary.cell(row=2, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

summary_data = [
    ('店铺中心', 14, 6, 3, 5),
    ('商品中心', 17, 13, 1, 3),
    ('价格中心', 15, 7, 5, 3),
    ('库存中心', 8, 7, 0, 1),
    ('内容中心', 23, 2, 1, 20),
    ('视频表现', 18, 9, 5, 4),
    ('流量中心', 14, 7, 5, 2),
    ('广告中心', 22, 15, 5, 2),
    ('订单中心', 17, 17, 0, 0),
    ('物流中心', 11, 7, 2, 2),
    ('评价中心', 11, 6, 2, 3),
    ('财务中心', 13, 7, 1, 5),
    ('事件中心', 18, 0, 7, 11),
]

for row_idx, (name, total, yes, warn, no) in enumerate(summary_data, 3):
    ws_summary.cell(row=row_idx, column=1, value=name).font = Font(name='Arial', bold=True, size=10, color='1E293B')
    ws_summary.cell(row=row_idx, column=1).border = thin_border
    ws_summary.cell(row=row_idx, column=1).alignment = left_align

    for col in range(2, 6):
        ws_summary.cell(row=row_idx, column=col).font = center_font
        ws_summary.cell(row=row_idx, column=col).border = thin_border
        ws_summary.cell(row=row_idx, column=col).alignment = center_align
    ws_summary.cell(row=row_idx, column=2, value=total)
    c_yes = ws_summary.cell(row=row_idx, column=3, value=yes)
    c_yes.font = yes_font; c_yes.fill = yes_fill
    c_warn = ws_summary.cell(row=row_idx, column=4, value=warn)
    c_warn.font = warn_font; c_warn.fill = warn_fill
    c_no = ws_summary.cell(row=row_idx, column=5, value=no)
    c_no.font = no_font; c_no.fill = no_fill

    pct = round(yes / total * 100) if total > 0 else 0
    ws_summary.cell(row=row_idx, column=6, value=f'{pct}%').font = Font(name='Arial', bold=True, size=10, color='2563EB')

    ws_summary.row_dimensions[row_idx].height = 24

total_row = len(summary_data) + 3
ws_summary.cell(row=total_row, column=1, value='合计').font = Font(name='Arial', bold=True, size=11, color='1E293B')
ws_summary.cell(row=total_row, column=1).border = thin_border
ws_summary.cell(row=total_row, column=1).alignment = left_align
for col in range(2, 6):
    total_val = sum(s[col-1] for s in summary_data)
    ws_summary.cell(row=total_row, column=col, value=total_val).font = Font(name='Arial', bold=True, size=10)
    ws_summary.cell(row=total_row, column=col).border = thin_border
    ws_summary.cell(row=total_row, column=col).alignment = center_align
total_all = sum(s[1] for s in summary_data)
total_yes = sum(s[2] for s in summary_data)
pct_total = round(total_yes / total_all * 100) if total_all > 0 else 0
ws_summary.cell(row=total_row, column=6, value=f'{pct_total}%').font = Font(name='Arial', bold=True, size=11, color='2563EB')
ws_summary.cell(row=total_row, column=6).border = thin_border
ws_summary.cell(row=total_row, column=6).alignment = center_align

wb.save('docs/虾掌柜_TikTok_SDK字段审计.xlsx')
print('Done! Saved to docs/虾掌柜_TikTok_SDK字段审计.xlsx')
